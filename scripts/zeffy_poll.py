#!/usr/bin/env -S uv run
# /// script
# requires-python = ">=3.11"
# dependencies = [
#     "requests>=2.32",
#     "openpyxl>=3.1",
#     "PyYAML>=6.0",
# ]
# ///
"""
One-shot poll of a Zeffy campaign's guest list.

Reads scripts/cookies.json (produced by zeffy_login.py), hits the Zeffy
internal export endpoint for the named campaign, decodes the returned XLSX,
and writes a Hugo data file in the form:

    last_updated: "2026-06-01 10:30 PDT"
    registrations:
      - full_name: ...
        ticket: 1
        nickname: ...

State and idempotency:
    A scripts/state.json file remembers the set of (ticket, full_name)
    keys we have already published.  If a poll's results match what's
    already in state, no file is written, no commit is made, no push
    is attempted — the run is a no-op.  When new registrations appear,
    the YAML is rewritten, a single commit is created, and (with --push)
    pushed.

Auth expiry:
    On HTTP 401/403, state.json gets cookies_expired=True and the
    process exits 0.  Run zeffy_login.py to refresh.

Concurrency:
    Uses fcntl.flock on scripts/.lock; a second invocation while one
    is already running exits 0 with a log message.

Usage:
    uv run scripts/zeffy_poll.py --campaign <UUID> --output <PATH>
    uv run scripts/zeffy_poll.py --campaign <UUID> --output <PATH> --push
    uv run scripts/zeffy_poll.py --campaign <UUID> --output <PATH> --dry-run
"""

import argparse
import base64
import fcntl
import io
import json
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import requests
import openpyxl
import yaml

ENDPOINT = "https://api.zeffy.com/_new/trpc/exportGuestList"

# Timestamp on the rendered page: Pacific time (PDT or PST per date).
DISPLAY_TZ = ZoneInfo("America/Los_Angeles")

SCRIPT_DIR = Path(__file__).parent
REPO_ROOT = SCRIPT_DIR.parent
COOKIES_PATH = SCRIPT_DIR / "cookies.json"
DEFAULT_STATE_PATH = SCRIPT_DIR / "state.json"
LOCK_PATH = SCRIPT_DIR / ".lock"


def build_body(campaign_id: str) -> dict:
    return {
        "occurrenceIds": [],
        "ticketingId": campaign_id,
        "filters": {
            "searchFilter": "",
            "statusFilter": [],
            "rateFilter": [],
            "answerFilter": [],
        },
        "exportParameters": {
            "locale": "EN",
            "timezone": "America/Los_Angeles",
        },
        "exportFileType": "Excel",
        "columnsSelection": {
            "selectedColumns": ["guestName", "buyerName", "ticketNumber",
                                "ticketType", "customQuestions"],
        },
    }


def load_session() -> requests.Session:
    if not COOKIES_PATH.exists():
        sys.exit(f"No cookies at {COOKIES_PATH}.  Run zeffy_login.py first.")
    session = requests.Session()
    for c in json.loads(COOKIES_PATH.read_text()):
        session.cookies.set(c["name"], c["value"], domain=c.get("domain"))
    return session


class AuthExpired(Exception):
    pass


# Browser-imitating headers.  Zeffy is fronted by Cloudflare; the
# `__cf_bm` bot-management cookie is partially bound to the client
# fingerprint that produced it.  Bare requests-default headers get
# served the Cloudflare "Just a moment..." JS challenge (HTTP 403,
# text/html body) instead of the export.  These headers mimic the
# Chrome that the login flow ran in, which is enough to satisfy CF.
_BROWSER_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)
_API_HEADERS = {
    "Content-Type": "application/json",
    "User-Agent": _BROWSER_UA,
    "Accept": "*/*",
    "Accept-Language": "en-US,en;q=0.9",
    # Don't advertise brotli — requests doesn't auto-decompress it.
    "Accept-Encoding": "gzip, deflate",
    "Origin": "https://www.zeffy.com",
    "Referer": "https://www.zeffy.com/en-US/o/fundraising/campaigns/hub",
    "sec-ch-ua": '"Not_A Brand";v="8", "Chromium";v="120", "Google Chrome";v="120"',
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": '"macOS"',
    "Sec-Fetch-Dest": "empty",
    "Sec-Fetch-Mode": "cors",
    "Sec-Fetch-Site": "same-site",
}


def fetch_xlsx_bytes(session: requests.Session, campaign_id: str) -> bytes:
    resp = session.post(
        ENDPOINT,
        json=build_body(campaign_id),
        headers=_API_HEADERS,
        timeout=30,
    )
    if resp.status_code in (401, 403):
        raise AuthExpired(f"HTTP {resp.status_code}")
    resp.raise_for_status()
    payload = resp.json()
    try:
        b64 = payload["result"]["data"]["data"]["export"]
    except (KeyError, TypeError):
        sys.exit(f"Unexpected response shape: {json.dumps(payload)[:500]}")
    return base64.b64decode(b64)


def parse_ticket_number(s: str) -> int | None:
    # Zeffy returns either bare digits ("13") or "#13 General Admission"
    # depending on the campaign's ticket-type setup.  Take the first run of
    # digits, optionally preceded by '#'.
    m = re.search(r"#?\s*(\d+)", s or "")
    return int(m.group(1)) if m else None


# Map Zeffy's per-ticket-type strings to abstract attendance flags so we
# can aggregate across an attendee's multiple tickets and produce a
# display label.  Adding a new ticket type means adding it here; unknown
# strings pass through verbatim in format_ticket_type_flags.
TICKET_TYPE_FLAGS: dict[str, set[str]] = {
    "BARGE 2026 (no Banquet)": {"BARGE"},
    "Banquet Only":            {"Banquet"},
    "BARGE 2026 + Banquet":    {"BARGE", "Banquet"},
}


def ticket_type_to_flags(raw: str) -> set[str]:
    """Map a raw Zeffy ticket-type to {"BARGE"} / {"Banquet"} / both."""
    if raw in TICKET_TYPE_FLAGS:
        return set(TICKET_TYPE_FLAGS[raw])
    # Unknown ticket type — pass the raw string through as a single flag
    # so format_ticket_type_flags surfaces it in the display.
    return {raw} if raw else set()


def format_ticket_type_flags(flags: set[str]) -> str:
    """Compose the display string from an attendee's aggregated flags."""
    if "BARGE" in flags and "Banquet" in flags:
        return "BARGE & Banquet"
    if "BARGE" in flags:
        return "BARGE"
    if "Banquet" in flags:
        return "Banquet"
    # Only unknown flags left — show them verbatim.
    return " & ".join(sorted(flags)) if flags else ""


def parse_workbook(xlsx_bytes: bytes, debug: bool = False) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if debug:
        print(f"  sheets: {wb.sheetnames}")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if debug:
        print(f"  raw row count: {len(rows)}")
        for i, r in enumerate(rows[:5]):
            print(f"  row[{i}]: {r}")
    if not rows:
        return []

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    if debug:
        print(f"  header: {header}")

    def find_col(*candidates: str) -> int | None:
        for cand in candidates:
            for i, h in enumerate(header):
                if cand in h:
                    return i
        return None

    # Column-preference order for the displayed name:
    #   "Name For Badge"  — custom question, used by campaigns where one
    #                       buyer can register multiple attendees (BARGE).
    #   "Guest name"      — Zeffy's built-in column (populated when each
    #                       ticket has its own guest profile; e.g. ATLARGE).
    #   "Buyer name"      — last-resort fallback.
    badge_col = find_col("name for badge", "badge name")
    guest_col = find_col("guest")
    buyer_col = find_col("buyer")
    ticket_col = find_col("ticket number")
    type_col = find_col("ticket type")
    notes_col = find_col("nickname", "questions", "notes")

    if ticket_col is None or (badge_col is None and guest_col is None
                              and buyer_col is None):
        sys.exit(f"Could not find name/ticket columns in header: {header}")

    def cell(row: tuple, idx: int | None) -> str:
        # openpyxl read-only mode truncates trailing Nones, so a row may be
        # shorter than the header.  Treat missing/None as empty.
        if idx is None or idx >= len(row) or row[idx] is None:
            return ""
        return str(row[idx]).strip()

    out: list[dict] = []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        full_name = (cell(r, badge_col) or cell(r, guest_col)
                     or cell(r, buyer_col))
        if not full_name:
            continue
        ticket_num = parse_ticket_number(cell(r, ticket_col))
        nickname = cell(r, notes_col)
        ticket_type_raw = cell(r, type_col)
        out.append({
            "full_name": full_name,
            "ticket": ticket_num,
            "nickname": nickname,
            "ticket_type_raw": ticket_type_raw,
        })

    out.sort(key=lambda r: (r["ticket"] is None, r["ticket"] or 0))

    # Dedup + aggregate ticket types.  A single buyer may purchase
    # multiple ticket types (e.g. BARGE registration + Banquet, or a
    # bundle) and appears as one XLSX row per ticket.  Collapse to one
    # entry per attendee, keyed by (Name For Badge, Nickname For Badge):
    # both are user-supplied custom-question values consistent across
    # the same person's purchases, and the pair is more unique than
    # full_name alone (which can collide).  After sorting by ticket
    # above, the first occurrence wins for the displayed ticket number,
    # yielding the smallest (earliest) ticket per attendee — a
    # reasonable "registration order" proxy.  Across the group we union
    # the ticket-type flags so e.g. someone with BARGE + Banquet shows
    # as "BARGE & Banquet" just like someone who bought the bundle.
    by_key: dict[tuple[str, str], dict] = {}
    flags_by_key: dict[tuple[str, str], set[str]] = {}
    for r in out:
        key = (r["full_name"], r["nickname"])
        if key not in by_key:
            by_key[key] = {
                "full_name": r["full_name"],
                "ticket": r["ticket"],
                "nickname": r["nickname"],
            }
            flags_by_key[key] = set()
        flags_by_key[key] |= ticket_type_to_flags(r["ticket_type_raw"])

    deduped: list[dict] = []
    for key, entry in by_key.items():
        entry["ticket_type"] = format_ticket_type_flags(flags_by_key[key])
        deduped.append(entry)
    return deduped


def key_for(row: dict) -> str:
    return f"{row['ticket']}|{row['full_name']}"


def load_state(state_path: Path, output_path: Path,
               campaign_id: str) -> dict:
    """Load state.json, or bootstrap published_keys from the existing
    YAML data file if state.json is absent.

    The bootstrap matters on a fresh checkout (or after state.json gets
    deleted): without it, the first run would treat every current
    registrant as "new" and produce a misleading "+N new" commit
    message even though only a few actually changed.
    """
    if state_path.exists():
        try:
            return json.loads(state_path.read_text())
        except json.JSONDecodeError as e:
            sys.exit(f"State file at {state_path} is not valid JSON: {e}")

    if not output_path.exists():
        return {}
    try:
        existing = yaml.safe_load(output_path.read_text())
    except yaml.YAMLError:
        return {}
    if isinstance(existing, dict):
        rows = existing.get("registrations") or []
    elif isinstance(existing, list):
        rows = existing
    else:
        rows = []
    keys = [f"{r.get('ticket')}|{r.get('full_name')}"
            for r in rows if r.get("full_name")]
    if not keys:
        return {}
    print(f"Bootstrapping state from existing {output_path} "
          f"({len(keys)} entries).")
    return {
        "campaign_id": campaign_id,
        "published_keys": sorted(keys),
    }


def save_state(path: Path, state: dict) -> None:
    path.write_text(json.dumps(state, indent=2, sort_keys=True) + "\n")


def acquire_lock():
    # Caller leaks the file handle until process exit, which releases the
    # OS-level lock.  Returns None if another process holds the lock.
    fd = open(LOCK_PATH, "w")
    try:
        fcntl.flock(fd.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
    except BlockingIOError:
        fd.close()
        return None
    return fd


def run_git(args: list[str]) -> subprocess.CompletedProcess:
    return subprocess.run(
        ["git"] + args, cwd=REPO_ROOT,
        capture_output=True, text=True,
    )


def commit_changes(output_path: Path, n_new: int) -> None:
    """Stage and commit the output file.  No push — the caller decides."""
    msg = (f"Zeffy: +{n_new} new registration"
           f"{'s' if n_new != 1 else ''} ({output_path.name})")
    add = run_git(["add", str(output_path.resolve())])
    if add.returncode != 0:
        sys.exit(f"git add failed: {add.stderr.strip()}")

    # If nothing was actually staged (e.g. state was wrong and the file
    # content matches HEAD), there's nothing to commit — recover quietly.
    staged = run_git(["diff", "--cached", "--quiet",
                      str(output_path.resolve())])
    if staged.returncode == 0:
        print("File unchanged from HEAD; skipping commit (state corrected).")
        return

    commit = run_git(["commit", "-m", msg])
    if commit.returncode != 0:
        sys.exit(f"git commit failed: {commit.stderr.strip()}")
    print(f"Committed: {msg}")


def push_changes() -> None:
    """Pull-rebase then push HEAD to origin.  Exits nonzero on failure.

    The rebase handles the common case where someone else (Tim, a deploy
    job, anyone) has pushed since our last sync.  Our commit only ever
    touches the YAML data file; outside changes essentially never touch
    that file (it's auto-managed), so the rebase succeeds cleanly almost
    always.  A conflict means someone manually edited the YAML on origin
    — we abort the rebase and exit loud rather than risk silently
    overwriting their work.

    Critically, the caller must save state.json BEFORE calling this — a
    failed push must not leave state thinking nothing was published.
    Otherwise every subsequent run will re-detect the same "new" entries,
    re-commit them (bumping just the timestamp), re-fail to push, and
    pile up hundreds of redundant local commits.  We learned this the
    hard way on 2026-06-03; don't undo the ordering."""
    pull = run_git(["pull", "--rebase", "origin", "main"])
    if pull.returncode != 0:
        # Make sure we don't leave the repo stuck mid-rebase.
        run_git(["rebase", "--abort"])
        sys.exit("git pull --rebase failed before push.  Investigate "
                 "manually — likely a conflict on the YAML data file.\n"
                 f"stderr: {pull.stderr.strip()}")

    pushres = run_git(["push"])
    if pushres.returncode != 0:
        sys.exit("git push failed.  Investigate manually — DO NOT force-push.\n"
                 f"stderr: {pushres.stderr.strip()}")
    print("Pushed.")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--campaign", required=True,
                    help="Zeffy campaign ID (the formId in the campaign URL).")
    ap.add_argument("--output", required=True, type=Path,
                    help="Path to the Hugo data YAML file to write.")
    ap.add_argument("--state", type=Path, default=DEFAULT_STATE_PATH,
                    help="Path to state.json (default: scripts/state.json).")
    ap.add_argument("--push", action="store_true",
                    help="After committing, also git push.  The watcher will "
                         "pass this; manual runs default to commit-only.")
    ap.add_argument("--dry-run", action="store_true",
                    help="Print parsed rows; no writes, no commits, no push.")
    args = ap.parse_args()

    lock = acquire_lock()
    if lock is None:
        print("Another zeffy_poll is running; exiting.")
        return 0

    state = load_state(args.state, args.output, args.campaign)

    session = load_session()
    try:
        xlsx_bytes = fetch_xlsx_bytes(session, args.campaign)
    except AuthExpired as e:
        print(f"Auth failed ({e}); marking cookies_expired and exiting.",
              file=sys.stderr)
        state["cookies_expired"] = True
        state["last_run_at"] = datetime.now(DISPLAY_TZ).isoformat()
        state["last_run_status"] = "auth_expired"
        save_state(args.state, state)
        return 0

    raw_path = SCRIPT_DIR / "last_export.xlsx"
    raw_path.write_bytes(xlsx_bytes)
    print(f"Wrote raw export to {raw_path} ({len(xlsx_bytes)} bytes).")

    rows = parse_workbook(xlsx_bytes, debug=args.dry_run)

    if not rows:
        sys.exit("Parsed 0 rows.  Refusing to clobber data file.  "
                 "Check cookies and campaign id.")

    current_keys = {key_for(r) for r in rows}
    # If the campaign id we're polling doesn't match the one in state, the
    # operator has switched events; treat published as empty so the whole
    # current list gets written and committed.
    if state.get("campaign_id") == args.campaign:
        published_keys = set(state.get("published_keys", []))
    else:
        published_keys = set()
    new_keys = current_keys - published_keys

    print(f"Parsed {len(rows)} row(s); "
          f"{len(new_keys)} new since last published.")

    if args.dry_run:
        for r in rows:
            marker = " *" if key_for(r) in new_keys else "  "
            print(f"  #{r['ticket']!s:>4}{marker} {r['full_name']!r:<40}  "
                  f"nick={r['nickname']!r}")
        return 0

    if not new_keys:
        # Nothing to publish; just record that we ran and exit clean.
        state["last_run_at"] = datetime.now(DISPLAY_TZ).isoformat()
        state["last_run_status"] = "no_changes"
        state["cookies_expired"] = False
        save_state(args.state, state)
        print("No new registrations; data file and state unchanged.")
        return 0

    # Order:
    #   1. Write the YAML.
    #   2. git add + commit (succeeds locally even if origin is ahead).
    #   3. Save state.json — state reflects what's now in local HEAD.
    #   4. THEN attempt the push.
    # Push failures must not leave state stale; see push_changes() docstring.
    payload = {
        "last_updated": datetime.now(DISPLAY_TZ).strftime("%Y-%m-%d %H:%M %Z"),
        "registrations": rows,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        yaml.safe_dump(payload, sort_keys=False, allow_unicode=True)
    )
    print(f"Wrote {args.output}")

    commit_changes(args.output, len(new_keys))

    state["campaign_id"] = args.campaign
    state["published_keys"] = sorted(current_keys)
    state["last_run_at"] = datetime.now(DISPLAY_TZ).isoformat()
    state["last_run_status"] = "ok"
    state["cookies_expired"] = False
    save_state(args.state, state)

    if args.push:
        push_changes()
    return 0


if __name__ == "__main__":
    sys.exit(main())
